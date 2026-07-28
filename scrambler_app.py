#!/usr/bin/env python3
"""Excel Scrambler — Fernet-based column scrambler/unscrambler for Excel files.

Scrambles selected columns of an .xlsx/.xlsm workbook with a freshly generated
Fernet key, saves the key (with metadata about the file, sheet, columns and
timestamp) into a local key library, and can later restore the original values
from any key in the library. Never modifies the input file: output is always a
new "<name>_scrambled.xlsx" / "<name>_unscrambled.xlsx" copy.
"""

import base64
import json
import os
import queue
import stat
import struct
import subprocess
import sys
import threading
import uuid
from datetime import datetime, date, time as dtime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from cryptography.fernet import Fernet, InvalidToken
from cryptography.hazmat.primitives import hashes, hmac as crypto_hmac
from cryptography.hazmat.primitives import padding as sym_padding
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

APP_NAME = "Excel Scrambler"
APP_DIR = Path(__file__).resolve().parent
KEY_DIR = Path.home() / ".excel-scrambler" / "keys"
TOKEN_PREFIX = "XSCRAMBLE1:"  # marks a cell as scrambled by this app
EXCEL_TYPES = [("Excel workbooks", "*.xlsx *.xlsm")]
SAMPLE_LEN = 24


# ---------------------------------------------------------------- key store

def ensure_key_dir():
    KEY_DIR.mkdir(parents=True, exist_ok=True)
    os.chmod(KEY_DIR.parent, stat.S_IRWXU)
    os.chmod(KEY_DIR, stat.S_IRWXU)


def save_key_record(record: dict) -> Path:
    ensure_key_dir()
    path = KEY_DIR / f"{record['id']}.json"
    with open(path, "w") as f:
        json.dump(record, f, indent=2)
    os.chmod(path, stat.S_IRUSR | stat.S_IWUSR)
    return path


def load_key_records() -> list:
    ensure_key_dir()
    records = []
    for p in KEY_DIR.glob("*.json"):
        try:
            with open(p) as f:
                rec = json.load(f)
            if rec.get("app") == "excel-scrambler" and rec.get("key"):
                rec["_path"] = str(p)
                records.append(rec)
        except (json.JSONDecodeError, OSError):
            continue
    records.sort(key=lambda r: r.get("created", ""), reverse=True)
    return records


# ------------------------------------------------------------ cell encoding

def encode_value(value):
    """Serialize a cell value (preserving its type) to a JSON payload."""
    if isinstance(value, bool):
        return json.dumps({"t": "bool", "v": value})
    if isinstance(value, int):
        return json.dumps({"t": "int", "v": value})
    if isinstance(value, float):
        return json.dumps({"t": "float", "v": value})
    if isinstance(value, datetime):
        return json.dumps({"t": "datetime", "v": value.isoformat()})
    if isinstance(value, date):
        return json.dumps({"t": "date", "v": value.isoformat()})
    if isinstance(value, dtime):
        return json.dumps({"t": "time", "v": value.isoformat()})
    return json.dumps({"t": "str", "v": str(value)})


def decode_value(payload: str):
    obj = json.loads(payload)
    t, v = obj["t"], obj["v"]
    if t == "datetime":
        return datetime.fromisoformat(v)
    if t == "date":
        return date.fromisoformat(v)
    if t == "time":
        return dtime.fromisoformat(v)
    return v


# --------------------------------------------------- deterministic encryption

def deterministic_encrypt(key_b64: bytes, data: bytes) -> bytes:
    """Fernet-compatible token whose IV is derived from the plaintext.

    Identical inputs produce byte-identical tokens (so equal cells stay
    matchable after scrambling); the token still decrypts with plain
    Fernet.decrypt(). Trade-off: equal values are visibly equal, which
    leaks repetition patterns — that's the point of the mode.
    """
    key = base64.urlsafe_b64decode(key_b64)
    signing_key, enc_key = key[:16], key[16:]

    h = crypto_hmac.HMAC(enc_key, hashes.SHA256())
    h.update(b"excel-scrambler-deterministic-iv:" + data)
    iv = h.finalize()[:16]

    padder = sym_padding.PKCS7(128).padder()
    padded = padder.update(data) + padder.finalize()
    enc = Cipher(algorithms.AES(enc_key), modes.CBC(iv)).encryptor()
    ct = enc.update(padded) + enc.finalize()

    # Fernet wire format: version | timestamp | IV | ciphertext | HMAC.
    # Timestamp is fixed at 0 so equal plaintexts yield equal tokens.
    basic = b"\x80" + struct.pack(">Q", 0) + iv + ct
    hm = crypto_hmac.HMAC(signing_key, hashes.SHA256())
    hm.update(basic)
    return base64.urlsafe_b64encode(basic + hm.finalize())


# ------------------------------------------------------------ core engine

def unique_output_path(src: Path, suffix: str) -> Path:
    base = src.with_name(f"{src.stem}_{suffix}{src.suffix}")
    if not base.exists():
        return base
    n = 2
    while True:
        cand = src.with_name(f"{src.stem}_{suffix} ({n}){src.suffix}")
        if not cand.exists():
            return cand
        n += 1


def scramble_file(src: Path, sheet: str, columns: list, skip_rows: int,
                  deterministic: bool = False, reuse_key: str = None):
    """Returns (output_path, key_record_path, cells_scrambled).

    reuse_key: an existing library key (base64 string) to scramble with
    instead of generating a fresh one — lets several files share one key.
    A separate record is still saved per file, each carrying the same key.
    """
    keep_vba = src.suffix.lower() == ".xlsm"
    wb = load_workbook(src, keep_vba=keep_vba)
    ws = wb[sheet]
    key = reuse_key.encode("ascii") if reuse_key else Fernet.generate_key()
    fernet = Fernet(key)  # also validates a reused key up front

    start = skip_rows + 1
    count = 0
    for col in columns:
        for row in range(start, ws.max_row + 1):
            cell = ws[f"{col}{row}"]
            if cell.value is None or cell.value == "":
                continue
            if isinstance(cell.value, str) and cell.value.startswith(TOKEN_PREFIX):
                continue  # already scrambled — don't double-wrap
            payload = encode_value(cell.value).encode("utf-8")
            if deterministic:
                token = deterministic_encrypt(key, payload)
            else:
                token = fernet.encrypt(payload)
            cell.value = TOKEN_PREFIX + token.decode("ascii")
            count += 1

    out = unique_output_path(src, "scrambled")
    wb.save(out)

    record = {
        "app": "excel-scrambler",
        "version": 2,
        "id": uuid.uuid4().hex[:12],
        "created": datetime.now().astimezone().isoformat(timespec="seconds"),
        "key": key.decode("ascii"),
        "source_file": str(src),
        "output_file": str(out),
        "sheet": sheet,
        "columns": columns,
        "skip_rows": skip_rows,
        "deterministic": deterministic,
        "reused_key": bool(reuse_key),
        "cells_scrambled": count,
    }
    key_path = save_key_record(record)
    return out, key_path, count


def record_skip_rows(rec: dict) -> int:
    """Rows-to-skip for a key record, honoring old header_exempt records."""
    if "skip_rows" in rec:
        return int(rec["skip_rows"])
    return 1 if rec.get("header_exempt", True) else 0


def unscramble_file(src: Path, sheet: str, columns: list, skip_rows: int,
                    key_b64: str):
    """Returns (output_path, cells_restored, failures, skipped_plain)."""
    keep_vba = src.suffix.lower() == ".xlsm"
    wb = load_workbook(src, keep_vba=keep_vba)
    ws = wb[sheet]
    fernet = Fernet(key_b64.encode("ascii"))

    start = skip_rows + 1
    restored = failures = skipped = 0
    for col in columns:
        for row in range(start, ws.max_row + 1):
            cell = ws[f"{col}{row}"]
            v = cell.value
            if v is None or v == "":
                continue
            if not (isinstance(v, str) and v.startswith(TOKEN_PREFIX)):
                skipped += 1
                continue
            token = v[len(TOKEN_PREFIX):].encode("ascii")
            try:
                payload = fernet.decrypt(token).decode("utf-8")
                cell.value = decode_value(payload)
                restored += 1
            except (InvalidToken, ValueError, json.JSONDecodeError):
                failures += 1

    if restored == 0 and failures > 0:
        raise InvalidToken(
            f"None of the {failures} scrambled cells could be decrypted — "
            "this is almost certainly the wrong key for this file.")

    out = unique_output_path(src, "unscrambled")
    wb.save(out)
    return out, restored, failures, skipped


def sheet_column_info(path: Path, sheet: str):
    """[(letter, header, sample), ...] for every column in the used range.

    All columns are listed (and therefore selectable) — header/sample text
    is just a preview where the first few rows happen to have content.
    """
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    max_col = ws.max_column or 1
    headers, samples = {}, {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, max_col=max_col)):
        for cell in row:
            if cell.value in (None, ""):
                continue
            letter = get_column_letter(cell.column)
            if i == 0:
                headers.setdefault(letter, str(cell.value))
            elif letter not in samples:
                samples[letter] = str(cell.value)
    wb.close()
    return [(get_column_letter(c),
             headers.get(get_column_letter(c), ""),
             samples.get(get_column_letter(c), ""))
            for c in range(1, max_col + 1)]


def sheet_names(path: Path):
    wb = load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def truncate(s: str, n: int = SAMPLE_LEN):
    s = s.replace("\n", " ")
    return s if len(s) <= n else s[: n - 1] + "…"


def reveal_in_finder(path: str):
    """Show the file in Finder / Explorer / the default file manager."""
    if sys.platform == "darwin":
        subprocess.run(["open", "-R", path], check=False)
    elif sys.platform == "win32":
        subprocess.run(["explorer", f"/select,{os.path.normpath(path)}"],
                       check=False)
    else:
        subprocess.run(["xdg-open", str(Path(path).parent)], check=False)


# ------------------------------------------------------------------- UI

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("760x640")
        self.minsize(660, 560)
        style = ttk.Style(self)
        if "aqua" in style.theme_names():
            style.theme_use("aqua")
        style.configure("Title.TLabel", font=("Helvetica", 22, "bold"))
        style.configure("Sub.TLabel", foreground="#666")
        style.configure("Big.TButton", font=("Helvetica", 15), padding=12)
        self._screen = None
        self.show(MenuScreen)

    def show(self, screen_cls, **kwargs):
        if self._screen is not None:
            self._screen.destroy()
        self._screen = screen_cls(self, **kwargs)
        self._screen.pack(fill="both", expand=True)


class MenuScreen(ttk.Frame):
    def __init__(self, app: App):
        super().__init__(app, padding=40)
        self.app = app
        ttk.Label(self, text="Excel Scrambler", style="Title.TLabel").pack(pady=(30, 6))
        ttk.Label(self, text="Encrypt or restore columns of an Excel workbook "
                             "with a saved Fernet key.",
                  style="Sub.TLabel").pack(pady=(0, 40))

        box = ttk.Frame(self)
        box.pack()
        ttk.Button(box, text="🔒  Scramble a file", style="Big.TButton", width=26,
                   command=lambda: self.pick(mode="scramble")).pack(pady=10)
        ttk.Button(box, text="🔓  Unscramble a file", style="Big.TButton", width=26,
                   command=lambda: self.pick(mode="unscramble")).pack(pady=10)
        ttk.Button(box, text="🗝  Key library", style="Big.TButton", width=26,
                   command=lambda: self.app.show(KeyLibraryScreen)).pack(pady=10)
        self.update_btn = ttk.Button(box, text="⟳  Check for updates",
                                     style="Big.TButton", width=26,
                                     command=self.update_app)
        self.update_btn.pack(pady=10)
        self.update_q = queue.Queue()

        n = len(load_key_records())
        ttk.Label(self, text=f"{n} key{'s' if n != 1 else ''} in library  ·  "
                             f"stored in {KEY_DIR}",
                  style="Sub.TLabel").pack(side="bottom", pady=10)

    def update_app(self):
        """Pull the latest version from GitHub and reinstall dependencies."""
        self.update_btn.state(["disabled"])
        self.update_btn.config(text="⟳  Updating…")

        def work():
            try:
                r = subprocess.run(["git", "pull", "--ff-only"], cwd=APP_DIR,
                                   capture_output=True, text=True, timeout=180)
                if r.returncode != 0:
                    self.update_q.put(("err", (r.stderr or r.stdout).strip()))
                    return
                if "Already up to date" in r.stdout:
                    self.update_q.put(("current", None))
                    return
                p = subprocess.run(
                    [sys.executable, "-m", "pip", "install", "--quiet",
                     "-r", str(APP_DIR / "requirements.txt")],
                    capture_output=True, text=True, timeout=600)
                if p.returncode != 0:
                    self.update_q.put(("err", "Code was updated but installing "
                                       "dependencies failed:\n"
                                       + (p.stderr or p.stdout).strip()))
                    return
                self.update_q.put(("updated", r.stdout.strip()))
            except FileNotFoundError:
                self.update_q.put(("err", "git was not found on this machine — "
                                   "run the update script in the app folder "
                                   "instead (update.sh / update.bat)."))
            except Exception as e:
                self.update_q.put(("err", str(e)))

        threading.Thread(target=work, daemon=True).start()
        self.after(150, self.poll_update)

    def poll_update(self):
        try:
            kind, detail = self.update_q.get_nowait()
        except queue.Empty:
            self.after(150, self.poll_update)
            return
        self.update_btn.state(["!disabled"])
        self.update_btn.config(text="⟳  Check for updates")
        if kind == "err":
            messagebox.showerror(APP_NAME, f"Update failed:\n\n{detail}",
                                 parent=self)
        elif kind == "current":
            messagebox.showinfo(APP_NAME, "You're already on the latest "
                                          "version.", parent=self)
        else:
            if messagebox.askyesno(
                    APP_NAME,
                    "Update installed.\n\nRestart Excel Scrambler now to use "
                    "the new version?", parent=self):
                os.execl(sys.executable, sys.executable,
                         str(APP_DIR / "scrambler_app.py"))

    def pick(self, mode):
        path = filedialog.askopenfilename(
            parent=self, title=f"Choose an Excel file to {mode}",
            filetypes=EXCEL_TYPES)
        if not path:
            return
        p = Path(path)
        if p.suffix.lower() not in (".xlsx", ".xlsm"):
            messagebox.showerror(APP_NAME, "Only Excel files (.xlsx / .xlsm) "
                                           "are accepted.", parent=self)
            return
        try:
            names = sheet_names(p)
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Could not open workbook:\n{e}",
                                 parent=self)
            return
        self.app.show(JobScreen, mode=mode, path=p, sheets=names)


class JobScreen(ttk.Frame):
    """Column/sheet/key selection and execution for both modes."""

    def __init__(self, app: App, mode: str, path: Path, sheets: list):
        super().__init__(app, padding=24)
        self.app, self.mode, self.path, self.sheets = app, mode, path, sheets
        self.busy = False
        self.result_q = queue.Queue()
        verb = "Scramble" if mode == "scramble" else "Unscramble"

        head = ttk.Frame(self)
        head.pack(fill="x")
        ttk.Button(head, text="‹ Back", command=self.go_back).pack(side="left")
        ttk.Label(head, text=f"{verb}: {path.name}",
                  font=("Helvetica", 16, "bold")).pack(side="left", padx=14)

        # sheet picker
        row = ttk.Frame(self)
        row.pack(fill="x", pady=(16, 8))
        ttk.Label(row, text="Sheet:").pack(side="left")
        self.sheet_var = tk.StringVar(value=sheets[0])
        combo = ttk.Combobox(row, textvariable=self.sheet_var, values=sheets,
                             state="readonly", width=28)
        combo.pack(side="left", padx=8)
        combo.bind("<<ComboboxSelected>>", lambda e: self.load_columns())

        ttk.Label(row, text="Skip first").pack(side="left", padx=(16, 4))
        self.skip_var = tk.IntVar(value=1)
        ttk.Spinbox(row, from_=0, to=9999, textvariable=self.skip_var,
                    width=5).pack(side="left")
        ttk.Label(row, text="row(s) — labels/headers stay untouched").pack(
            side="left", padx=4)

        if mode == "scramble":
            optrow = ttk.Frame(self)
            optrow.pack(fill="x", pady=(0, 8))
            self.det_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(
                optrow,
                text="Match-preserving mode — identical values scramble to the "
                     "identical token, so other programs can still match them "
                     "(reveals which cells are equal)",
                variable=self.det_var).pack(anchor="w")

            keyrow = ttk.Frame(self)
            keyrow.pack(fill="x", pady=(0, 8))
            self.keys = load_key_records()
            self.keymode_var = tk.StringVar(value="new")
            ttk.Radiobutton(keyrow, text="Generate a new key",
                            variable=self.keymode_var, value="new",
                            command=self.on_keymode).pack(side="left")
            ttk.Radiobutton(keyrow, text="Reuse an existing key:",
                            variable=self.keymode_var, value="reuse",
                            command=self.on_keymode).pack(side="left", padx=(16, 4))
            labels = [
                f"{Path(r.get('source_file', '?')).name} · "
                f"{r.get('created', '')[:16].replace('T', ' ')} · "
                f"cols {', '.join(r.get('columns', []))}"
                for r in self.keys]
            self.reuse_combo = ttk.Combobox(keyrow, values=labels,
                                            state="disabled", width=44)
            self.reuse_combo.pack(side="left", fill="x", expand=True)
            if labels:
                self.reuse_combo.current(0)

        # column checklist
        ttk.Label(self, text=f"Columns to {verb.lower()}:").pack(anchor="w")
        colwrap = ttk.Frame(self, relief="sunken", borderwidth=1)
        colwrap.pack(fill="both", expand=True, pady=(4, 10))
        self.canvas = tk.Canvas(colwrap, highlightthickness=0)
        vsb = ttk.Scrollbar(colwrap, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.col_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.col_frame, anchor="nw")
        self.col_frame.bind("<Configure>", lambda e: self.canvas.configure(
            scrollregion=self.canvas.bbox("all")))
        self.col_vars = {}

        # key picker (unscramble only)
        if mode == "unscramble":
            ttk.Label(self, text="Key from your library:").pack(anchor="w")
            keywrap = ttk.Frame(self)
            keywrap.pack(fill="x", pady=(4, 10))
            cols = ("file", "when", "sheet", "columns")
            self.keys = load_key_records()
            self.tree = ttk.Treeview(keywrap, columns=cols, show="headings",
                                     height=min(6, max(3, len(self.keys))))
            for cid, txt, w in (("file", "Original file", 240),
                                ("when", "Scrambled at", 160),
                                ("sheet", "Sheet", 90),
                                ("columns", "Columns", 90)):
                self.tree.heading(cid, text=txt)
                self.tree.column(cid, width=w, anchor="w")
            ksb = ttk.Scrollbar(keywrap, orient="vertical",
                                command=self.tree.yview)
            self.tree.configure(yscrollcommand=ksb.set)
            ksb.pack(side="right", fill="y")
            self.tree.pack(side="left", fill="x", expand=True)
            for rec in self.keys:
                when = rec.get("created", "")[:16].replace("T", "  ")
                self.tree.insert("", "end", iid=rec["id"], values=(
                    Path(rec.get("source_file", "?")).name, when,
                    rec.get("sheet", ""), ", ".join(rec.get("columns", []))))
            self.tree.bind("<<TreeviewSelect>>", self.on_key_selected)
            if not self.keys:
                ttk.Label(self, text="No keys in your library yet — scramble "
                                     "a file first.", style="Sub.TLabel").pack(anchor="w")
            else:
                self.preselect_key()

        # action row
        foot = ttk.Frame(self)
        foot.pack(fill="x", pady=(6, 0))
        self.status = ttk.Label(foot, text="", style="Sub.TLabel")
        self.status.pack(side="left")
        self.go_btn = ttk.Button(foot, text=f"{verb} ▸", style="Big.TButton",
                                 command=self.run)
        self.go_btn.pack(side="right")
        self.progress = ttk.Progressbar(foot, mode="indeterminate", length=160)

        self.load_columns()

    # -- helpers

    def go_back(self):
        if not self.busy:
            self.app.show(MenuScreen)

    def on_keymode(self):
        state = "readonly" if self.keymode_var.get() == "reuse" else "disabled"
        self.reuse_combo.config(state=state)

    def load_columns(self):
        for w in self.col_frame.winfo_children():
            w.destroy()
        self.col_vars = {}
        try:
            info = sheet_column_info(self.path, self.sheet_var.get())
        except Exception as e:
            messagebox.showerror(APP_NAME, f"Could not read sheet:\n{e}",
                                 parent=self)
            return
        if not info:
            ttk.Label(self.col_frame, text="(sheet is empty)").pack(anchor="w")
            return
        for letter, header, sample in info:
            var = tk.BooleanVar(value=False)
            desc = f"{letter}"
            if header:
                desc += f"   —   {truncate(header)}"
            if sample:
                desc += f"    (e.g. {truncate(sample)})"
            ttk.Checkbutton(self.col_frame, text=desc, variable=var).pack(
                anchor="w", padx=8, pady=2)
            self.col_vars[letter] = var

    def preselect_key(self):
        """Auto-select the newest key whose output file matches the chosen file."""
        target = str(self.path)
        for rec in self.keys:
            if rec.get("output_file") == target or \
                    Path(rec.get("output_file", "")).name == self.path.name:
                self.tree.selection_set(rec["id"])
                self.tree.see(rec["id"])
                return

    def on_key_selected(self, _evt=None):
        rec = self.selected_key()
        if not rec:
            return
        # mirror the key's recorded sheet, columns and header setting into the form
        key_sheet = rec.get("sheet")
        if key_sheet and key_sheet != self.sheet_var.get() and key_sheet in self.sheets:
            self.sheet_var.set(key_sheet)
            self.load_columns()
        for letter, var in self.col_vars.items():
            var.set(letter in rec.get("columns", []))
        self.skip_var.set(record_skip_rows(rec))

    def selected_key(self):
        if self.mode != "unscramble":
            return None
        sel = self.tree.selection()
        if not sel:
            return None
        return next((r for r in self.keys if r["id"] == sel[0]), None)

    # -- execution

    def run(self):
        if self.busy:
            return
        columns = [c for c, v in self.col_vars.items() if v.get()]
        if not columns:
            messagebox.showwarning(APP_NAME, "Select at least one column.",
                                   parent=self)
            return
        rec = None
        if self.mode == "unscramble":
            rec = self.selected_key()
            if not rec:
                messagebox.showwarning(APP_NAME, "Select a key from the library.",
                                       parent=self)
                return

        reuse_key = None
        if self.mode == "scramble" and self.keymode_var.get() == "reuse":
            idx = self.reuse_combo.current()
            if idx < 0 or idx >= len(self.keys):
                messagebox.showwarning(
                    APP_NAME, "Pick which existing key to reuse (or switch "
                              "back to 'Generate a new key').", parent=self)
                return
            reuse_key = self.keys[idx]["key"]

        self.busy = True
        self.go_btn.state(["disabled"])
        self.progress.pack(side="right", padx=10)
        self.progress.start(12)
        self.status.config(text="Working…")

        sheet = self.sheet_var.get()
        try:
            skip_rows = max(0, int(self.skip_var.get()))
        except (tk.TclError, ValueError):
            messagebox.showwarning(APP_NAME, "Rows to skip must be a number.",
                                   parent=self)
            self.busy = False
            self.go_btn.state(["!disabled"])
            self.progress.stop()
            self.progress.pack_forget()
            self.status.config(text="")
            return
        deterministic = bool(self.det_var.get()) if self.mode == "scramble" else False
        path = self.path

        def work():
            try:
                if self.mode == "scramble":
                    out, key_path, n = scramble_file(path, sheet, columns,
                                                     skip_rows, deterministic,
                                                     reuse_key)
                    self.result_q.put(("ok-scramble", out, key_path, n))
                else:
                    out, ok, bad, skipped = unscramble_file(
                        path, sheet, columns, skip_rows, rec["key"])
                    self.result_q.put(("ok-unscramble", out, ok, bad, skipped))
            except Exception as e:
                self.result_q.put(("err", str(e)))

        threading.Thread(target=work, daemon=True).start()
        self.after(120, self.poll)

    def poll(self):
        try:
            res = self.result_q.get_nowait()
        except queue.Empty:
            self.after(120, self.poll)
            return
        self.busy = False
        self.progress.stop()
        self.progress.pack_forget()
        self.go_btn.state(["!disabled"])
        self.status.config(text="")

        kind = res[0]
        if kind == "err":
            messagebox.showerror(APP_NAME, f"Operation failed:\n\n{res[1]}",
                                 parent=self)
            return
        if kind == "ok-scramble":
            _, out, key_path, n = res
            messagebox.showinfo(APP_NAME,
                f"Scrambled {n} cells.\n\nNew file:\n{out}\n\n"
                f"Key saved to your library:\n{key_path}", parent=self)
        else:
            _, out, ok, bad, skipped = res
            msg = f"Restored {ok} cells.\n\nNew file:\n{out}"
            if bad:
                msg += f"\n\n⚠️ {bad} cells could not be decrypted with this key."
            if skipped:
                msg += f"\n({skipped} cells were not scrambled and were left as-is.)"
            messagebox.showinfo(APP_NAME, msg, parent=self)
        reveal_in_finder(str(res[1]))
        self.app.show(MenuScreen)


class KeyLibraryScreen(ttk.Frame):
    def __init__(self, app: App):
        super().__init__(app, padding=24)
        self.app = app
        head = ttk.Frame(self)
        head.pack(fill="x")
        ttk.Button(head, text="‹ Back",
                   command=lambda: app.show(MenuScreen)).pack(side="left")
        ttk.Label(head, text="Key library",
                  font=("Helvetica", 16, "bold")).pack(side="left", padx=14)

        cols = ("file", "when", "sheet", "columns", "cells")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for cid, txt, w in (("file", "Original file", 230),
                            ("when", "Scrambled at", 150),
                            ("sheet", "Sheet", 90),
                            ("columns", "Columns", 90),
                            ("cells", "Cells", 60)):
            self.tree.heading(cid, text=txt)
            self.tree.column(cid, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, pady=12)
        self.refresh()

        foot = ttk.Frame(self)
        foot.pack(fill="x")
        ttk.Button(foot, text="Show scrambled file in Finder",
                   command=self.reveal_file).pack(side="left")
        ttk.Button(foot, text="Delete key…",
                   command=self.delete_key).pack(side="right")

    def refresh(self):
        self.tree.delete(*self.tree.get_children())
        self.keys = load_key_records()
        for rec in self.keys:
            when = rec.get("created", "")[:16].replace("T", "  ")
            self.tree.insert("", "end", iid=rec["id"], values=(
                Path(rec.get("source_file", "?")).name, when,
                rec.get("sheet", ""), ", ".join(rec.get("columns", [])),
                rec.get("cells_scrambled", "")))

    def _selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning(APP_NAME, "Select a key first.", parent=self)
            return None
        return next((r for r in self.keys if r["id"] == sel[0]), None)

    def reveal_file(self):
        rec = self._selected()
        if rec:
            out = rec.get("output_file", "")
            if out and Path(out).exists():
                reveal_in_finder(out)
            else:
                messagebox.showinfo(APP_NAME, "The scrambled file no longer "
                                              "exists at its original location.",
                                    parent=self)

    def delete_key(self):
        rec = self._selected()
        if not rec:
            return
        if not messagebox.askyesno(
                APP_NAME,
                "Deleting this key is PERMANENT.\n\nAny file scrambled with it "
                "can never be unscrambled again.\n\nDelete anyway?",
                icon="warning", default="no", parent=self):
            return
        try:
            os.remove(rec["_path"])
        except OSError as e:
            messagebox.showerror(APP_NAME, f"Could not delete key file:\n{e}",
                                 parent=self)
            return
        self.refresh()


def main():
    ensure_key_dir()
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
